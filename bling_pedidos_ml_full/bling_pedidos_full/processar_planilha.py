"""
processar_planilha.py  —  Robô da Conferência ML Full.

Agrupa os SKUs em blocos separados por LINHA EM BRANCO (cada bloco = 1 pedido),
reserva o estoque no depósito "Estoque Geral" e devolve a planilha preenchida:
    Coluna B (Nº PEDIDO DE VENDA) = número final do pedido criado
    Coluna H (QNT PEDIDO DE VENDA) = quantidade efetivamente reservada

Pode ser usado pela interface gráfica (gui.py) ou pelo terminal:
    python processar_planilha.py --arquivo entrada.xlsx --dry-run
    python processar_planilha.py --arquivo entrada.xlsx
"""

import os
import sys
import csv
import argparse
import datetime as dt

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
from bling_client import BlingClient, BlingError

FILL_FALTA = PatternFill("solid", fgColor="FFE2A8")  # laranja para faltas

ABA = "Montagem de Pedidos - Envio"
COL_SKU = 4   # D
COL_QTD = 7   # G  (planejada)
COL_PED = 2   # B  (saída: número do pedido)
COL_RES = 8   # H  (saída: quantidade reservada)
LINHA_INICIO = 4


class ProcessoErro(Exception):
    pass


def detectar_blocos(ws):
    blocos, atual = [], []
    for r in range(LINHA_INICIO, ws.max_row + 1):
        sku = ws.cell(r, COL_SKU).value
        if sku in (None, "") or str(sku).strip() == "":
            if atual:
                blocos.append(atual); atual = []
        else:
            atual.append(r)
    if atual:
        blocos.append(atual)
    return blocos


def ler_item(ws, r):
    sku = str(ws.cell(r, COL_SKU).value).strip()
    try:
        planejada = float(ws.cell(r, COL_QTD).value or 0)
    except (TypeError, ValueError):
        planejada = 0.0
    return sku, planejada


def processar(arquivo, dry_run=False, sem_estoque=False, saldo_fisico=False,
              data_venda=None, prazo_dias=None, obs_interna="", emit=print):
    """Processa a planilha. `emit(texto)` recebe cada linha de progresso.
    data_venda: date (padrão hoje); prazo_dias: dias p/ data prevista (padrão 3);
    obs_interna: texto que vai nas observações internas de todos os pedidos.
    Retorna um dict com resultados. Lança ProcessoErro em caso de problema."""
    load_dotenv()

    if data_venda is None:
        data_venda = dt.date.today()
    if prazo_dias is None:
        prazo_dias = int(os.environ.get("BLING_PRAZO_DIAS", "3"))
    data_prevista = data_venda + dt.timedelta(days=int(prazo_dias))

    wb = load_workbook(arquivo)  # preserva fórmulas e formatação
    if ABA not in wb.sheetnames:
        raise ProcessoErro(f"Aba '{ABA}' não encontrada. Abas: {wb.sheetnames}")
    ws = wb[ABA]
    blocos = detectar_blocos(ws)
    total_plan = sum(ler_item(ws, r)[1] for b in blocos for r in b)

    emit("=" * 60)
    emit(f"Arquivo : {os.path.basename(arquivo)}")
    emit(f"Blocos  : {len(blocos)}  ->  até {len(blocos)} pedidos de venda")
    emit(f"Pares   : {int(total_plan)} planejados no total")
    emit(f"Modo    : {'SIMULAÇÃO (não cria pedidos)' if dry_run else 'CRIAÇÃO REAL'}")
    emit(f"Data venda: {data_venda:%d/%m/%Y}  |  Prevista: {data_prevista:%d/%m/%Y}  (+{int(prazo_dias)} dias)")
    if obs_interna:
        emit(f"Obs. interna: {obs_interna}")
    emit("=" * 60)

    if dry_run:
        for i, b in enumerate(blocos, 1):
            itens = [ler_item(ws, r) for r in b]
            pares = int(sum(q for _, q in itens))
            emit(f"  Pedido {i:>2} (linhas {b[0]}-{b[-1]}): {len(itens)} SKUs / {pares} pares")
        emit("\n(simulação) Nada foi enviado ao Bling.")
        return {"dry_run": True, "blocos": len(blocos)}

    contato_id = os.environ.get("BLING_CONTATO_ID")
    if not contato_id:
        raise ProcessoErro("Cliente fake não definido (BLING_CONTATO_ID). Configure na interface.")
    for var in ("BLING_CLIENT_ID", "BLING_CLIENT_SECRET", "BLING_REDIRECT_URI"):
        if not os.environ.get(var):
            raise ProcessoErro(f"Credencial faltando: {var}. Configure e autorize o Bling primeiro.")

    client = BlingClient(
        client_id=os.environ["BLING_CLIENT_ID"],
        client_secret=os.environ["BLING_CLIENT_SECRET"],
        redirect_uri=os.environ["BLING_REDIRECT_URI"],
    )

    usar_virtual = not saldo_fisico
    id_dep = None
    if not sem_estoque:
        nome_dep = os.environ.get("BLING_DEPOSITO", "Estoque Geral")
        id_dep = client.achar_deposito_id(nome_dep)
        if not id_dep:
            raise ProcessoErro(f"Depósito '{nome_dep}' não encontrado no Bling. Confira o nome.")
        emit(f"Depósito '{nome_dep}' -> id {id_dep}")

    obs_prefixo = os.environ.get("BLING_OBS_PREFIXO", "Reserva ML Full")
    valor_padrao = float(os.environ.get("BLING_VALOR_PADRAO", "0"))

    log, faltas, pedidos_com_falta = [], [], []
    emit("\nCriando pedidos...")
    for i, bloco in enumerate(blocos, 1):
        itens_payload, linhas = [], {}
        for r in bloco:
            sku, planejada = ler_item(ws, r)
            pid = client.resolver_produto_id(sku)
            if not pid:
                linhas[r] = {"sku": sku, "plan": planejada, "res": 0, "motivo": "produto não encontrado"}
                continue
            if sem_estoque:
                reservavel = planejada
            else:
                saldo = client.saldo_no_deposito(pid, id_dep, usar_virtual)
                reservavel = max(0, min(planejada, saldo))
            motivo = "" if reservavel >= planejada else ("sem saldo" if reservavel == 0 else "saldo parcial")
            linhas[r] = {"sku": sku, "plan": planejada, "res": reservavel, "motivo": motivo}
            if reservavel > 0:
                itens_payload.append({"produto": {"id": pid}, "codigo": sku,
                                      "quantidade": reservavel, "valor": valor_padrao})

        numero, status, erro = "", "", ""
        if itens_payload:
            payload = {"contato": {"id": int(contato_id)},
                       "data": data_venda.isoformat(),
                       "dataSaida": data_venda.isoformat(),
                       "dataPrevista": data_prevista.isoformat(),
                       "observacoesInternas": obs_interna,
                       "itens": itens_payload}
            sit = os.environ.get("BLING_SITUACAO_ID")
            if sit:
                payload["situacao"] = {"id": int(sit)}
            resp = client.criar_pedido_venda(payload)
            if resp.status_code in (200, 201):
                numero = resp.json().get("data", {}).get("numero", "")
                status = "OK"
            else:
                status, erro = f"ERRO {resp.status_code}", resp.text[:300]
                emit(f"  Pedido {i:>2} FALHOU: {resp.status_code} {resp.text[:140]}")
        else:
            status = "SEM_SALDO"

        falta_skus, total_faltou = 0, 0
        for r in bloco:
            info = linhas[r]
            ws.cell(r, COL_PED).value = numero
            ws.cell(r, COL_RES).value = info["res"]
            if info["res"] < info["plan"]:
                ws.cell(r, COL_RES).fill = FILL_FALTA
                falta_skus += 1
                total_faltou += info["plan"] - info["res"]
                faltas.append([i, numero, info["sku"], int(info["plan"]),
                               int(info["res"]), int(info["plan"] - info["res"]), info["motivo"]])

        reservado = int(sum(linhas[r]["res"] for r in bloco))
        if status == "OK":
            marca = "  ATENCAO: faltando itens" if falta_skus else ""
            emit(f"  Pedido {i:>2} (linhas {bloco[0]}-{bloco[-1]}) -> nº {numero} | {reservado} pares{marca}")
        elif status == "SEM_SALDO":
            emit(f"  Pedido {i:>2} (linhas {bloco[0]}-{bloco[-1]}): SEM SALDO — nenhum pedido criado")
        if falta_skus or status == "SEM_SALDO":
            pedidos_com_falta.append((i, numero or "(não criado)", falta_skus or len(bloco), int(total_faltou)))
        log.append([i, f"{bloco[0]}-{bloco[-1]}", numero, status, erro])

    pasta = os.path.dirname(os.path.abspath(arquivo))
    base = os.path.splitext(os.path.basename(arquivo))[0]
    ts = f"{dt.datetime.now():%Y%m%d_%H%M%S}"
    saida = os.path.join(pasta, f"{base}_PREENCHIDA_{ts}.xlsx")
    wb.save(saida)

    rel = os.path.join(pasta, f"log_{ts}.csv")
    with open(rel, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f); w.writerow(["bloco", "linhas", "numero_pedido", "status", "erro"]); w.writerows(log)

    rel_faltas = ""
    if faltas:
        rel_faltas = os.path.join(pasta, f"faltas_{ts}.csv")
        with open(rel_faltas, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["bloco", "numero_pedido", "sku", "planejada", "reservada", "faltou", "motivo"])
            w.writerows(faltas)

    ok = sum(1 for x in log if x[3] == "OK")
    emit(f"\nConcluído: {ok}/{len(blocos)} pedidos criados.")
    emit(f"Planilha preenchida salva (células laranja na coluna H = reserva abaixo do planejado).")

    if pedidos_com_falta:
        emit(f"\nATENCAO: {len(pedidos_com_falta)} pedido(s) com itens faltando:")
        for bloco, num, qtd_skus, faltou in pedidos_com_falta:
            emit(f"   Pedido nº {num} (bloco {bloco}): {qtd_skus} SKU(s) abaixo do planejado, faltaram {faltou} pares")
    else:
        emit("\nTodos os pedidos reservaram a quantidade planejada integralmente.")

    return {"dry_run": False, "saida": saida, "log": rel, "faltas": rel_faltas,
            "pedidos_ok": ok, "blocos": len(blocos), "pedidos_com_falta": len(pedidos_com_falta)}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--arquivo", required=True)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--sem-estoque", action="store_true")
    ap.add_argument("--saldo-fisico", action="store_true")
    ap.add_argument("--obs", default="", help="observação interna (nº do Full do mês)")
    ap.add_argument("--prazo", type=int, default=None, help="dias para a data prevista (padrão 3)")
    args = ap.parse_args()
    try:
        r = processar(args.arquivo, dry_run=args.dry_run, sem_estoque=args.sem_estoque,
                      saldo_fisico=args.saldo_fisico, prazo_dias=args.prazo,
                      obs_interna=args.obs, emit=print)
        if not r.get("dry_run"):
            print(f"\nArquivos gerados em: {os.path.dirname(os.path.abspath(args.arquivo))}")
    except (ProcessoErro, BlingError) as e:
        print(f"[ERRO] {e}"); sys.exit(1)


if __name__ == "__main__":
    main()
